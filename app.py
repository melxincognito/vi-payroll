from flask import Flask, render_template, request, send_file, jsonify
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import uuid
import threading
import time

app = Flask(__name__)

REQUIRED_COLUMNS = [
    "Enrollment ID", "VI Teacher", "School or District Name", "Section",
    "Base Section Name", "Tier", "Amount", "Quarter Pay", "Student First Name",
    "Student Last Name", "Creation Date", "First Activity Date", "Last Activity Date",
    "Start Date", "Due Date", "Days Active", "Expected Progress", "Actual Progress",
    "Course Grade", "Completable Items", "Completed Items", "Total Minutes",
    "Messages Sent", "Publisher", "UserSpace"
]

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ALT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
TOTAL_FONT = Font(color="FFFFFF", bold=True, size=11)
BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

# In-memory job store: { job_id: { progress, message, status, output, error } }
jobs = {}


def style_sheet(ws, df):
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill

    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(col_name)),
            *[len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(2, ws.max_row + 1)]
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def add_totals_row(ws, df):
    cols = list(df.columns)
    total_row = ws.max_row + 1
    amount_col = cols.index("Amount") + 1 if "Amount" in cols else None
    quarter_pay_col = cols.index("Quarter Pay") + 1 if "Quarter Pay" in cols else None

    for col_idx in range(1, len(cols) + 1):
        cell = ws.cell(row=total_row, column=col_idx)
        if col_idx == 1:
            cell.value = "TOTAL"
        elif col_idx == amount_col:
            cell.value = df["Amount"].sum()
            cell.number_format = '"$"#,##0.00'
        elif col_idx == quarter_pay_col:
            cell.value = df["Quarter Pay"].sum()
            cell.number_format = '"$"#,##0.00'
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def safe_sheet_name(name, existing_names):
    name = str(name).strip()
    for ch in ['\\', '/', '*', '?', ':', '[', ']']:
        name = name.replace(ch, ' ')
    name = name[:31]
    base = name
    count = 1
    while name in existing_names:
        suffix = f" ({count})"
        name = base[:31 - len(suffix)] + suffix
        count += 1
    return name


def run_job(job_id, file_bytes):
    def update(progress, message):
        jobs[job_id]["progress"] = progress
        jobs[job_id]["message"] = message

    try:
        update(5, "Reading file...")
        df = pd.read_excel(io.BytesIO(file_bytes))

        update(15, "Validating columns...")
        col_map = {c.lower().strip(): c for c in df.columns}
        matched = {}
        missing = []
        for req in REQUIRED_COLUMNS:
            key = req.lower().strip()
            if key in col_map:
                matched[req] = col_map[key]
            else:
                missing.append(req)

        if missing:
            jobs[job_id]["status"] = "error"
            jobs[job_id]["error"] = f"Missing columns: {', '.join(missing)}"
            return

        update(20, "Organizing columns...")
        df = df[[matched[c] for c in REQUIRED_COLUMNS]]
        df.columns = REQUIRED_COLUMNS

        teachers = df["VI Teacher"].dropna().unique()
        teachers_sorted = sorted(teachers, key=lambda x: str(x).strip().lower())
        total_teachers = len(teachers_sorted)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            update(25, "Writing All Students tab...")
            df.to_excel(writer, sheet_name="All Students", index=False)
            ws = writer.sheets["All Students"]
            style_sheet(ws, df)

            existing_names = {"All Students"}
            for i, teacher in enumerate(teachers_sorted):
                pct = 30 + int((i / total_teachers) * 65)
                update(pct, f"Processing {teacher} ({i + 1} of {total_teachers})...")

                teacher_df = df[df["VI Teacher"] == teacher].reset_index(drop=True)
                sheet_name = safe_sheet_name(teacher, existing_names)
                existing_names.add(sheet_name)
                teacher_df.to_excel(writer, sheet_name=sheet_name, index=False)
                ws = writer.sheets[sheet_name]
                style_sheet(ws, teacher_df)
                add_totals_row(ws, teacher_df)

        update(98, "Saving file...")
        output.seek(0)
        jobs[job_id]["output"] = output.read()
        update(100, "Done!")
        jobs[job_id]["status"] = "done"

        # Clean up after 5 minutes
        def cleanup():
            time.sleep(300)
            jobs.pop(job_id, None)
        threading.Thread(target=cleanup, daemon=True).start()

    except Exception as e:
        jobs[job_id]["status"] = "error"
        jobs[job_id]["error"] = str(e)


@app.route("/", methods=["GET"])
def index():
    return render_template("index.html")


@app.route("/process", methods=["POST"])
def process():
    if "file" not in request.files or request.files["file"].filename == "":
        return jsonify({"error": "Please upload an Excel file."}), 400

    file = request.files["file"]
    if not file.filename.endswith((".xlsx", ".xls")):
        return jsonify({"error": "Only .xlsx or .xls files are supported."}), 400

    file_bytes = file.read()
    job_id = str(uuid.uuid4())
    jobs[job_id] = {"progress": 0, "message": "Starting...", "status": "processing", "output": None, "error": None}

    thread = threading.Thread(target=run_job, args=(job_id, file_bytes), daemon=True)
    thread.start()

    return jsonify({"job_id": job_id})


@app.route("/status/<job_id>")
def status(job_id):
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job not found."}), 404
    return jsonify({
        "status": job["status"],
        "progress": job["progress"],
        "message": job["message"],
        "error": job.get("error"),
    })


@app.route("/download/<job_id>")
def download(job_id):
    job = jobs.get(job_id)
    if not job or job["status"] != "done":
        return "File not ready.", 404
    return send_file(
        io.BytesIO(job["output"]),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="VI_Teacher_Payroll.xlsx"
    )


if __name__ == "__main__":
    import os
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
